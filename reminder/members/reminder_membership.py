from flask import *
from flask import Blueprint, render_template, request, jsonify, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from .. models.membersModel import Member
from .. function import convert_phone_number, sendFile_wa_GAPI, sendFile_wa_GAPI_group, send_wa_GreenAPI
from datetime import date, datetime, timedelta
from extensions import db
from flask_recaptcha import ReCaptcha # Import ReCaptcha object
import pytz, random, json, jwt
import requests
import openpyxl
import xlrd
import pandas as pd
from random import randint
from time import sleep
from bs4 import BeautifulSoup
from decouple import config

reminder_membership = Blueprint("reminder_membership", __name__, static_folder="static", template_folder="templates")

base_url = config('BASE_URL')
current_time = datetime.now(pytz.timezone('Africa/Lusaka'))  
project_enviroment  = config('ENVIROMENT')
recaptcha = ReCaptcha()


@reminder_membership.route("/registration", methods=['POST'])
def registration():

    _json = request.json

    _platform = _json['platform']
    _firstname = _json['firstname']
    _lastname = _json['lastname']
    _wanumber = _json['wanumber']
    _dob = _json['dob']
    _accepted_terms = _json['accepted_terms']

    # Checking if values has been passed from the client
    if _platform and _firstname and _lastname and _wanumber and _dob and _accepted_terms:

        # Check if mobile is already in the database
        is_user_mobile = Member.query.filter_by(wa_number = convert_phone_number(_wanumber)).first()

        if is_user_mobile:
            resp = jsonify({'status': 400,
                            'isError': 'true',  
                            'message' : 'Mobile Number has already been used to register an account.'}), 400
            return resp
        else:

            if request.method == 'POST': # Check to see if flask.request.method is POST
                if recaptcha.verify(): # Use verify() method to see if ReCaptcha is filled out

                        created_date = current_time.strftime('%Y-%m-%d %H:%M:%S')
                        token = jwt.encode({'id': _wanumber, 'exp' : datetime.utcnow() + timedelta(minutes=60)}, config('SECRET_KEY'), algorithm='HS256') 

                        if _accepted_terms != "False":

                            members = Member(firstname = _firstname, lastname = _lastname, wa_number = convert_phone_number(_wanumber), dob= _dob, accepted_terms = True, created_at = created_date)
                            db.session.add(members)
                            db.session.commit()

                            resp = jsonify({'status': 200,
                                            'isError': 'false',
                                            'message' : registered_members}), 200
                            return resp

                        else:
                            members = Member(firstname = _firstname, lastname = _lastname, wa_number = convert_phone_number(_wanumber), dob= _dob, accepted_terms = False, created_at = created_date)
                            db.session.add(members)
                            db.session.commit()

                            resp = jsonify({'status': 200,
                                            'isError': 'false',
                                            'message' : registered_members}), 200
                            return resp
                    
                else:
                    resp = jsonify({'status': 400,
                            'isError': 'true',
                            'message' : 'Please fill out the ReCaptcha!'}), 400
                    return resp
            else:
                    resp = jsonify({'status': 400,
                            'isError': 'true',
                            'message' : 'Method not allowed'}), 400
                    return resp

     # If the keys are not passed
    else:
        resp = jsonify({'status': 400,
                        'isError': 'true',
                        'message' : 'One or more key values are missing, please enter missing values'}), 400
        return resp


#MLFC SERMON NOTES
@reminder_membership.route("/download_notes", methods=['POST'])
def download_notes():

        # Get data from the database
        members = Member.query.all()

        registered_members = []
        
        for person in members:
            person_details = {}
            person_details['memberid']= person.memberid
            person_details['firstname']= person.firstname
            person_details['lastname']= person.lastname
            person_details['wa_number']= person.wa_number
            person_details['dob'] = person.dob

            registered_members.append(person_details)

        url = 'https://mlfc.org/ministries/connection-groups/'
        reqs = requests.get(url)
        soup = BeautifulSoup(reqs.text, 'html.parser')
        
        i = 0
        for link in soup.find_all('a'):
            if ('-CG-Notes' in link.get('href')):
                i += 1
            
                # Get response object for link
                response = requests.get(link.get('href'))
        
                # Write content in pdf file
                pdf = open("pdf_upload/"+str(i)+".pdf", 'wb')
                pdf.write(response.content)
                pdf.close()
                print("File ", i, " downloaded")

        for x in registered_members:
            res = sendFile_wa_GAPI("260" + x['wa_number'], link.get('href'), "Hi " +  x['firstname'] + " " + x['lastname'] +" please find sermon study notes for this week. Please read through and let us meet on wednesday")
            sleep(randint(3,12))

        # load the json to a string
        JsonString = json.loads(res)

        resp = jsonify({'status': 200,
                        'isError': 'false',
                        'message' : registered_members}), 200
        return resp

#HAPPY BIRTHDAY MESSAGE FROM THE CG
@reminder_membership.route("/happy_birthday", methods=['POST'])
def happy_birthday():

            current_date = current_time.strftime('%Y-%m-%d')
   
            # Get data from the database
            members = Member.query.filter_by(dob = current_date)

            registered_members = []
            
            for person in members:
                person_details = {}
                person_details['memberid']= person.memberid
                person_details['firstname']= person.firstname
                person_details['lastname']= person.lastname
                person_details['wa_number']= person.wa_number
                person_details['dob'] = person.dob

                registered_members.append(person_details)

            for x in registered_members:
                res = send_wa_GreenAPI("260" + x['wa_number'],  "Hi " +  x['firstname'] + " " + x['lastname'] +". You are special to us and you are special to God. We as MLFC Carwash CG would like to wish you a happy birthday. We pray that you have a great one blessings")
                sleep(randint(3,12))
        
            resp = jsonify({'status': 200,
                            'isError': 'false',
                            'message' : registered_members}), 200
            return resp

        


@reminder_membership.route("/reading_reminder", methods=['POST'])
def reading_reminder():

   
            # Get data from the database
            members = Member.query.all()

            registered_members = []
            
            for person in members:
                person_details = {}
                person_details['memberid']= person.memberid
                person_details['firstname']= person.firstname
                person_details['lastname']= person.lastname
                person_details['wa_number']= person.wa_number
                person_details['dob'] = person.dob

                registered_members.append(person_details)

            # Define variable to load the dataframe
            dataframe = openpyxl.load_workbook("pdf_upload/ReadingPlan.xlsx", data_only=True)
            
            # Define variable to read sheet
            dataframe1 = dataframe.active

            second_column = dataframe1['B']
             # Create the list
            verses = [cell.value for cell in second_column[0:]]

            dt = date.today()
            rplan = ""
            
            
            # Iterate the loop to read the cell values
            for row in range(0, dataframe1.max_row):
                for col in dataframe1.iter_cols(1, dataframe1.max_column):
                    if col[row].value == datetime.combine(dt, datetime.min.time()):
                        
                        rplan = verses[row]
            

            for x in registered_members:
                res = send_wa_GreenAPI("260" + x['wa_number'],  "Hi " +  x['firstname'] + " " + x['lastname'] +" Today's " + str(dt)  + " bible reading plan is from " + rplan)
                sleep(randint(3,12))
            
            resp = jsonify({'status': 200,
                            'isError': 'false',
                            'message' : registered_members}), 200
            return resp





